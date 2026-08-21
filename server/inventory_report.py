"""
inventory_report.py

Builds the "Current Inventory" Excel report — a snapshot of everything
checked in via Incoming Inventory and its tagged location, plus a
count-by-location summary. Generated fresh on each request (not a template
someone edits and re-uploads), so plain values are used throughout rather
than formulas — there's nothing here that needs to recalculate.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import inventory

HEADER_FILL = PatternFill(start_color="1E232B", end_color="1E232B", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16)
BODY_FONT = Font(name="Arial", size=10)


def build_report():
    entries = inventory.list_entries(include_removed=False)

    wb = Workbook()

    # --- Sheet 1: Current Inventory ---
    ws = wb.active
    ws.title = "Current Inventory"

    ws["A1"] = "AES Logistics — Current Inventory"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

    headers = ["Job Number", "Location", "Confirmed By", "Confirmed At", "Photo Filename", "Note"]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    row = header_row + 1
    for entry in entries:
        ws.cell(row=row, column=1, value=entry["job_number"]).font = BODY_FONT
        ws.cell(row=row, column=2, value=entry["location"]).font = BODY_FONT
        ws.cell(row=row, column=3, value=entry["confirmed_by"]).font = BODY_FONT
        ws.cell(row=row, column=4, value=entry["confirmed_at"][:19].replace("T", " ")).font = BODY_FONT
        ws.cell(row=row, column=5, value=entry.get("photo_filename") or "").font = BODY_FONT
        ws.cell(row=row, column=6, value=entry.get("note") or "").font = BODY_FONT
        row += 1

    if not entries:
        ws.cell(row=row, column=1, value="(No inventory checked in yet)").font = Font(name="Arial", italic=True, color="999999")

    widths = [16, 14, 18, 20, 30, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{header_row + 1}"

    # --- Sheet 2: Summary by Location ---
    ws2 = wb.create_sheet("Summary by Location")
    ws2["A1"] = "Items Currently Checked In, by Location"
    ws2["A1"].font = TITLE_FONT

    ws2.cell(row=3, column=1, value="Location").font = HEADER_FONT
    ws2.cell(row=3, column=1).fill = HEADER_FILL
    ws2.cell(row=3, column=2, value="Count").font = HEADER_FONT
    ws2.cell(row=3, column=2).fill = HEADER_FILL

    counts = {loc: 0 for loc in inventory.LOCATIONS}
    for entry in entries:
        counts[entry["location"]] = counts.get(entry["location"], 0) + 1

    r = 4
    for loc in inventory.LOCATIONS:
        ws2.cell(row=r, column=1, value=loc).font = BODY_FONT
        ws2.cell(row=r, column=2, value=counts.get(loc, 0)).font = BODY_FONT
        r += 1

    ws2.cell(row=r, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
    ws2.cell(row=r, column=2, value=len(entries)).font = Font(name="Arial", bold=True)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
